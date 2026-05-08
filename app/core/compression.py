import json
import logging

import duckdb

from app.core.study_utils import list_study_tables
from app.core.config import settings

import os

logger = logging.getLogger(__name__)

SCHEMA_CACHE: dict[str, dict] = {}
ORIGINAL_TOKENS_CACHE: dict[str, int] = {}
TARGET_TOKEN_BUDGET = 3500


def count_tokens(text: str) -> float:
    return len(text.split()) * 1.3

def get_original_tokens(study_id: str, db_path: str = settings.DB_PATH) -> int:
    if study_id in ORIGINAL_TOKENS_CACHE:
        return ORIGINAL_TOKENS_CACHE[study_id]
        
    try:
        con = duckdb.connect(db_path)
        # First, try to load from DB
        row = con.execute("SELECT original_tokens FROM studies WHERE id = ?", (study_id,)).fetchone()
        if row and row[0] is not None:
            con.close()
            ORIGINAL_TOKENS_CACHE[study_id] = row[0]
            return row[0]

        # Fallback: compute it
        paths = con.execute("SELECT storage_path FROM files WHERE study_id = ? AND file_type = 'SDTM_CSV'", (study_id,)).fetchall()
        con.close()
        
        total = 0
        for (path,) in paths:
            if path and os.path.exists(path):
                total += int(os.path.getsize(path) / 4)
                
        # Persist to DB
        try:
            con = duckdb.connect(db_path)
            con.execute("UPDATE studies SET original_tokens = ? WHERE id = ?", (total, study_id))
            con.close()
        except Exception as e:
            logger.warning(f"Failed to persist original tokens for {study_id} to DB: {e}")

        ORIGINAL_TOKENS_CACHE[study_id] = total
        return total
    except Exception as e:
        logger.warning(f"Failed to calculate original tokens for {study_id}: {e}")
        return 0

def _extract_valid_values_from_schema(study_id: str, db_path: str) -> dict[str, dict[str, str]]:
    try:
        from openpyxl import load_workbook
        con = duckdb.connect(db_path)
        schema_file = con.execute(
            "SELECT storage_path FROM files WHERE file_type = 'Schema_JSON' AND study_id = ? ORDER BY created_at DESC",
            (study_id,)
        ).fetchone()
        con.close()

        if not schema_file:
            return {}

        path = schema_file[0]
        wb = load_workbook(path, read_only=True, data_only=True)
        mapping = {}
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            var_idx = -1
            val_idx = -1
            sheet_mapping = {}
            for row in sheet.iter_rows(values_only=True):
                if not row: continue
                if var_idx == -1:
                    for i, cell in enumerate(row):
                        val = str(cell).strip().lower() if cell else ''
                        if val == 'variable name': var_idx = i
                        elif val == 'valid values': val_idx = i
                elif var_idx != -1 and val_idx != -1:
                    var_name = row[var_idx]
                    val_vals = row[val_idx]
                    if var_name and val_vals and str(val_vals).strip():
                        val_str = str(val_vals).strip()
                        if len(val_str) > 200:
                            val_str = val_str[:197] + "..."
                        sheet_mapping[str(var_name).upper()] = val_str
            if sheet_mapping:
                mapping[sheet_name.upper()] = sheet_mapping
        return mapping
    except Exception as e:
        logger.warning(f"Could not extract valid values from Excel for study {study_id}: {e}")
        return {}


def _is_noise_column(column_name: str) -> bool:
    upper_name = column_name.upper()
    return (
        upper_name.endswith("_SEQ")
        or upper_name.endswith("_NUM")
        or upper_name == "STUDYID"
    )


def _is_priority_column(column_name: str) -> bool:
    upper_name = column_name.upper()
    return (
        "SUBJID" in upper_name
        or "USUBJID" in upper_name
        or "DT" in upper_name
        or upper_name.endswith("GR")
        or upper_name.endswith("SER")
        or upper_name.endswith("STRESN")
        or upper_name.endswith("STRESC")
    )


def compress_schema(
    study_id: str,
    db_path: str,
    created_tables: list[str] | None = None,
) -> dict:
    con = duckdb.connect(db_path, read_only=True)
    table_names = created_tables or list_study_tables(con, study_id)

    study_schema = {"study_id": study_id, "tables": {}}
    remaining_budget = TARGET_TOKEN_BUDGET
    valid_values_mapping = _extract_valid_values_from_schema(study_id, db_path)

    for table_name in table_names:
        columns_data = con.execute(f"DESCRIBE {table_name}").fetchall()
        priority_columns: list[dict[str, str]] = []
        standard_columns: list[dict[str, str]] = []

        for column_name, column_type, *_ in columns_data:
            if _is_noise_column(column_name):
                continue

            column_payload = {"name": column_name, "type": column_type}
            
            # Find matching valid values
            for sheet_name, sheet_mapping in valid_values_mapping.items():
                if table_name.upper().endswith(sheet_name.upper()):
                    if column_name.upper() in sheet_mapping:
                        column_payload["valid_values"] = sheet_mapping[column_name.upper()]
                    break

            if _is_priority_column(column_name):
                priority_columns.append(column_payload)
            else:
                standard_columns.append(column_payload)

        selected_columns = list(priority_columns)
        for column in standard_columns:
            tentative_columns = selected_columns + [column]
            tentative_schema = {
                "study_id": study_id,
                "tables": {
                    **study_schema["tables"],
                    table_name: {"columns": tentative_columns},
                },
            }
            if count_tokens(json.dumps(tentative_schema)) > TARGET_TOKEN_BUDGET:
                break
            selected_columns.append(column)

        study_schema["tables"][table_name] = {"columns": selected_columns}
        remaining_budget = TARGET_TOKEN_BUDGET - int(
            count_tokens(json.dumps(study_schema))
        )

    con.close()

    # Persist to DB
    try:
        con = duckdb.connect(db_path)
        con.execute("UPDATE studies SET compressed_schema = ? WHERE id = ?", (json.dumps(study_schema), study_id))
        con.close()
    except Exception as e:
        logger.warning(f"Failed to persist compressed schema for {study_id} to DB: {e}")

    SCHEMA_CACHE[study_id] = study_schema
    logger.info(
        "Compressed schema cached.",
        extra={
            "event_action": "schema_compression",
            "model_version": "none",
            "metadata": {
                "study_id": study_id,
                "tables": len(study_schema["tables"]),
                "estimated_tokens": int(count_tokens(json.dumps(study_schema))),
                "remaining_budget": remaining_budget,
            },
        },
    )
    return study_schema


def get_cached_schema(study_id: str, db_path: str = settings.DB_PATH) -> dict | None:
    if study_id in SCHEMA_CACHE:
        return SCHEMA_CACHE[study_id]
        
    # Lazy load from DB
    try:
        con = duckdb.connect(db_path, read_only=True)
        row = con.execute("SELECT compressed_schema FROM studies WHERE id = ?", (study_id,)).fetchone()
        con.close()
        if row and row[0]:
            schema = json.loads(row[0]) if isinstance(row[0], str) else row[0]
            SCHEMA_CACHE[study_id] = schema
            return schema
    except Exception as e:
        logger.error(f"Error lazy-loading schema for {study_id}: {e}")
        
    return None


def clear_cached_schema(study_id: str, db_path: str = settings.DB_PATH) -> None:
    SCHEMA_CACHE.pop(study_id, None)
    try:
        con = duckdb.connect(db_path)
        con.execute("UPDATE studies SET compressed_schema = NULL WHERE id = ?", (study_id,))
        con.close()
    except Exception as e:
        logger.warning(f"Failed to clear schema from DB for {study_id}: {e}")


def warm_schema_cache(db_path: str) -> None:
    """Called once at startup to populate the fast in-memory cache directly from DuckDB 
    instead of rebuilding the schema, enabling instant startup times."""
    try:
        con = duckdb.connect(db_path, read_only=True)
        rows = con.execute("SELECT id, compressed_schema, original_tokens FROM studies").fetchall()
        con.close()
        
        warmed = 0
        for study_id, schema_str, original_tokens in rows:
            if schema_str:
                schema = json.loads(schema_str) if isinstance(schema_str, str) else schema_str
                SCHEMA_CACHE[study_id] = schema
                warmed += 1
            if original_tokens is not None:
                ORIGINAL_TOKENS_CACHE[study_id] = original_tokens
                
        logger.info(
            "Schema cache loaded from persistent DB on startup.",
            extra={
                "event_action": "schema_compression",
                "model_version": "none",
                "metadata": {"studies_warmed": warmed},
            },
        )
    except Exception as exc:
        logger.warning(
            "warm_schema_cache: could not load studies from DB.",
            extra={
                "event_action": "schema_compression",
                "model_version": "none",
                "metadata": {"error": str(exc)},
            },
        )
