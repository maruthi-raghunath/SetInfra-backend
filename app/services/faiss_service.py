import json
import logging
import os
import gc
from typing import Any

# We move heavy imports (faiss, PyPDF2, etc.) inside functions to save RAM on startup.

logger = logging.getLogger(__name__)

def _get_gemini_client():
    from google import genai
    return genai.Client(api_key=settings.GEMINI_API_KEY)

def get_embeddings(texts: list[str]) -> Any:
    """Wrapper that chooses between local and remote embeddings based on settings."""
    from app.core.config import settings
    import numpy as np

    if settings.USE_LOCAL_EMBEDDING:
        logger.info(f"Computing local embeddings for {len(texts)} chunks...")
        from sentence_transformers import SentenceTransformer
        model = SentenceTransformer("all-MiniLM-L6-v2")
        embeddings = model.encode(texts, convert_to_numpy=True)
        # Clean up model immediately
        del model
        gc.collect()
        return np.array(embeddings, dtype=np.float32)
    else:
        logger.info(f"Fetching remote Gemini embeddings for {len(texts)} chunks...")
        try:
            client = _get_gemini_client()
            result = client.models.embed_content(
                model="text-embedding-004",
                contents=texts,
                config={"task_type": "RETRIEVAL_DOCUMENT"}
            )
            vectors = [e.values for e in result.embeddings]
            return np.array(vectors, dtype=np.float32)
        except Exception as e:
            logger.error(f"Gemini Remote Embedding failed: {e}")
            raise

def get_index_path(study_id: str) -> str:
    from app.core.config import settings
    return os.path.join(settings.VECTOR_DIR, f"{study_id}.index")

def get_chunks_path(study_id: str) -> str:
    from app.core.config import settings
    return os.path.join(settings.VECTOR_DIR, f"{study_id}.chunks.json")

def extract_text_from_file(file_path: str, file_type: str) -> str:
    if file_type == "Protocol":
        if file_path.lower().endswith(".pdf"):
            import PyPDF2
            text_parts: list[str] = []
            try:
                with open(file_path, "rb") as file_handle:
                    reader = PyPDF2.PdfReader(file_handle)
                    for page in reader.pages:
                        extracted = page.extract_text()
                        if extracted:
                            text_parts.append(extracted)
                return "\n".join(text_parts)
            except Exception as e:
                logger.error(f"PDF extraction failed: {e}")
                return ""

        if file_path.lower().endswith(".docx"):
            from docx import Document
            document = Document(file_path)
            return "\n".join(
                paragraph.text for paragraph in document.paragraphs if paragraph.text
            )

    if file_type == "Schema_JSON":
        if file_path.lower().endswith(".csv"):
            import csv
            with open(file_path, "r", encoding="utf-8", newline="") as file_handle:
                reader = csv.reader(file_handle)
                return "\n".join(",".join(cell for cell in row) for row in reader)

        if file_path.lower().endswith(".xlsx"):
            from openpyxl import load_workbook
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            sheet_text: list[str] = []
            for worksheet in workbook.worksheets:
                sheet_text.append(f"[Sheet: {worksheet.title}]")
                for row in worksheet.iter_rows(values_only=True):
                    values = ["" if value is None else str(value) for value in row]
                    if any(values):
                        sheet_text.append(",".join(values))
            workbook.close()
            return "\n".join(sheet_text)

    return ""

def chunk_text(text: str, chunk_size: int = 400, overlap: int = 50) -> list[str]:
    words = text.split()
    if not words:
        return []
    chunks: list[str] = []
    step = max(1, chunk_size - overlap)
    for start in range(0, len(words), step):
        chunk = " ".join(words[start : start + chunk_size]).strip()
        if chunk:
            chunks.append(chunk)
    return chunks

def index_documents(study_id: str, file_paths_with_types: list[tuple[str, str]]) -> str | None:
    import faiss
    from app.core.config import settings
    
    all_chunks: list[str] = []
    index = None
    batch_size = 50

    for file_path, file_type in file_paths_with_types:
        try:
            extracted_text = extract_text_from_file(file_path, file_type)
            if not extracted_text:
                continue
            file_chunks = chunk_text(extracted_text)
            del extracted_text
            gc.collect()

            if not file_chunks:
                continue

            for i in range(0, len(file_chunks), batch_size):
                batch = file_chunks[i : i + batch_size]
                normalized_embeddings = get_embeddings(batch)

                if index is None:
                    index = faiss.IndexFlatL2(normalized_embeddings.shape[1])
                index.add(normalized_embeddings)
            
            all_chunks.extend(file_chunks)
            gc.collect()
        except Exception as exc:
            logger.error(f"Failed to process file {file_path}: {exc}")
            continue

    if not all_chunks or index is None:
        return None

    os.makedirs(settings.VECTOR_DIR, exist_ok=True)
    index_path = get_index_path(study_id)
    faiss.write_index(index, index_path)
    
    chunks_path = get_chunks_path(study_id)
    with open(chunks_path, "w", encoding="utf-8") as f:
        json.dump({"study_id": study_id, "chunks": all_chunks}, f)

    # Cleanup
    del index
    gc.collect()

    logger.info(f"Index saved with {len(all_chunks)} chunks.")
    return index_path
